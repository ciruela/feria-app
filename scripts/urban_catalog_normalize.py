#!/usr/bin/env python3
"""Normaliza el Excel de catálogo de Urban Tactical (4 pestañas heterogéneas)
a un Excel limpio y canónico que el importador de la app entiende, incluyendo
las columnas de precios fijos (fiel al Excel, la app NO recalcula nada).

Pestañas de entrada:
  - Gral      : armas (pistolas/revólver = arma_corta; carabinas/escopetas/
                fusiles/PCC/DEC = arma_larga). Efectivo/PVP en USD, cuotas en ARS.
  - Taurus    : idem Gral pero columnas corridas (precio en E, descripción en F).
  - BERSA     : pistolas agrupadas por calibre, con variantes por acabado. ARS.
  - Accesorios: todo -> tipo accesorios. Cajas 3DURBAN (ARS) + fundas/linternas/
                cargadores/etc. (USD). Urban no llevó munición a la feria.

Salida: un único sheet "Catalogo" con encabezados canónicos + precios fijos.

Uso:
  python3 scripts/urban_catalog_normalize.py \
      ~/Downloads/Catalogo_Armas_Nuevas_Urban_con_Cantidad.xlsx \
      ~/Downloads/urban_catalog_clean.xlsx
"""
import sys
import re
import shutil
import zipfile

import openpyxl

TC = 1570.0  # tipo de cambio del Excel (celda R2 "TC:"); se lee dinámicamente en main()

OUT_HEADERS = [
    "tipo", "marca", "calibre", "modelo", "codigo", "descripcion",
    "precio_usd", "stock", "stock_inicial",
    "efectivo_ars", "efectivo_usd", "tarjeta_ars",
    "cuota3_ars", "cuota6_ars", "cuota12_ars",
]


def num(v):
    """Convierte una celda a float o None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = re.sub(r"[^\d.,\-]", "", s)
    if not s or s in {"-", ".", ","}:
        return None
    # Formato AR: 1.234,56 -> 1234.56 ; 1,50 -> 1.50 ; 1234.56 -> 1234.56
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def cell(ws, r, c):
    return ws.cell(row=r, column=c).value


def txt(v):
    if v is None:
        return ""
    return str(v).replace("\xa0", " ").strip()


def clean_desc(s):
    """Compacta la descripción (colapsa saltos/espacios múltiples)."""
    s = (s or "").replace("\xa0", " ")
    s = re.sub(r"\n{2,}", "\n", s)
    s = re.sub(r"[ \t]{2,}", " ", s)
    return s.strip()


def guess_calibre(text):
    """Extrae un calibre del nombre/desc cuando la columna Calibre viene vacía."""
    t = (text or "").replace("\xa0", " ")
    m = re.search(
        r"(\.\d{2,3}(?:\s*(?:LR|Special|Win|Mag|ACP|SPRG|S&W|SW|WMR|Rem))?)",
        t, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    m = re.search(r"\b(9\s*mm|5\.56|7\.62|12/70|20/76|\.?22\s*LR)\b", t, re.IGNORECASE)
    return m.group(1).strip() if m else ""


def arma_type(subcat):
    s = (subcat or "").upper()
    if "PISTOLA" in s or "REVOLVER" in s or "REVÓLVER" in s:
        return "arma_corta"
    return "arma_larga"


def rows_gral_taurus(wb, sheet, desc_col, price_lista_col):
    """Gral y Taurus. desc_col: columna descripción (5=E Gral, 6=F Taurus)."""
    ws = wb[sheet]
    out = []
    for r in range(3, ws.max_row + 1):
        sku = txt(cell(ws, r, 1))
        marca = txt(cell(ws, r, 2))
        nombre = txt(cell(ws, r, 3))
        if not sku and not nombre:
            continue
        calibre = txt(cell(ws, r, 4))
        descripcion = clean_desc(txt(cell(ws, r, desc_col)))
        subcat = txt(cell(ws, r, 7))
        stock = num(cell(ws, r, 8))
        efectivo_usd = num(cell(ws, r, 9))     # I
        pvp_tarjeta_usd = num(cell(ws, r, 10))  # J
        cuota3 = num(cell(ws, r, 11))           # K (ARS por cuota)
        cuota6 = num(cell(ws, r, 12))           # L
        cuota12 = num(cell(ws, r, 13))          # M
        if efectivo_usd is None:
            continue
        if not calibre:
            calibre = guess_calibre(nombre) or guess_calibre(descripcion)
        efectivo_ars = round(efectivo_usd * TC, 2)
        tarjeta_ars = round(pvp_tarjeta_usd * TC, 2) if pvp_tarjeta_usd else None
        out.append({
            "tipo": arma_type(subcat),
            "marca": marca or "Urban",
            "calibre": calibre,
            "modelo": nombre,
            "codigo": sku,
            "descripcion": descripcion or nombre,
            "precio_usd": round(efectivo_usd, 2),
            "stock": int(stock) if stock is not None else None,
            "stock_inicial": int(stock) if stock is not None else None,
            "efectivo_ars": efectivo_ars,
            "efectivo_usd": round(efectivo_usd, 2),
            "tarjeta_ars": tarjeta_ars,
            "cuota3_ars": cuota3,
            "cuota6_ars": cuota6,
            "cuota12_ars": cuota12,
        })
    return out


def bersa_calibre(model, section_cal):
    """Corrige el calibre de BERSA cuando la armería agrupó modelos bajo una
    sección equivocada. En el Excel no hay secciones ".40" ni ".45": los TPR40 y
    TPR45 quedan bajo "CALIBRE 9x19 MM". El nombre del modelo sí trae el calibre
    (TPR40, TPR45), así que lo derivamos de ahí. Si no hay match, se respeta la
    sección (fiel al Excel)."""
    m = (model or "").upper()
    if re.search(r"(?<!\d)45(?!\d)", m):
        return ".45 ACP"
    if re.search(r"(?<!\d)40(?!\d)", m):
        return ".40 S&W"
    return section_cal


def rows_bersa(wb):
    ws = wb["BERSA"]
    out = []
    current_cal = ""
    current_model = ""
    current_desc = ""
    for r in range(1, ws.max_row + 1):
        a = txt(cell(ws, r, 1))
        d = txt(cell(ws, r, 4))
        e = txt(cell(ws, r, 5))
        f = txt(cell(ws, r, 6))
        h = num(cell(ws, r, 8))     # precio sugerido (ARS)
        b = clean_desc(txt(cell(ws, r, 2)))
        # Sección de calibre
        m = re.match(r"CALIBRE\s+(.+)", a, re.IGNORECASE)
        if m:
            current_cal = m.group(1).strip()
            continue
        # Header de modelo: A tiene nombre y la fila trae rótulos (ACABADO/CÓDIGO)
        is_header = (d.upper() == "ACABADO" or e.upper() == "ACABADO"
                     or txt(cell(ws, r, 7)).upper() == "PRECIO")
        if a and is_header:
            current_model = a.strip()
            continue
        # Fila de variante: tiene código (F) y precio numérico (H)
        if f and h is not None:
            acabado = d if d and d.upper() != "ACABADO" else e
            if b:
                current_desc = b
            efectivo_ars = num(cell(ws, r, 10))  # J
            cuota3 = num(cell(ws, r, 11))
            cuota6 = num(cell(ws, r, 12))
            cuota12 = num(cell(ws, r, 13))
            modelo = (current_model + (f" {acabado}" if acabado else "")).strip()
            out.append({
                "tipo": "arma_corta",
                "marca": "BERSA",
                "calibre": bersa_calibre(current_model, current_cal),
                "modelo": modelo or current_model,
                "codigo": f,
                "descripcion": current_desc or modelo,
                "precio_usd": 0,
                "stock": int(num(cell(ws, r, 9))) if num(cell(ws, r, 9)) is not None else None,
                "stock_inicial": int(num(cell(ws, r, 9))) if num(cell(ws, r, 9)) is not None else None,
                "efectivo_ars": efectivo_ars if efectivo_ars else round(h, 2),
                "efectivo_usd": None,
                "tarjeta_ars": round(h, 2),
                "cuota3_ars": cuota3,
                "cuota6_ars": cuota6,
                "cuota12_ars": cuota12,
            })
    return out


def parse_accesorio_name(name):
    """('cajas de 9 mm x 100 municiones') -> (calibre, rounds_per_box)."""
    n = name.lower().replace("\xa0", " ")
    rounds = None
    mr = re.search(r"x\s*(\d{1,4})", n)
    if mr:
        rounds = int(mr.group(1))
    cal = ""
    # Token de calibre justo antes de "x N" (ej "9 mm x 100", "308 x 50").
    mc = re.search(r"([0-9]+(?:[./][0-9]+)?\s*(?:mm|l\.?)?)\s*[xX]\s*\d", n)
    if mc:
        cal = mc.group(1).strip().rstrip(".")
    return cal, rounds


def rows_accesorios(wb):
    """Hoja "Accesorios". TODO sale como tipo `accesorios` (Urban no llevó
    munición a la feria). La única diferencia es la moneda del precio:

    - Cajas 3DURBAN (cotizan en ARS) -> tipo `accesorios`, precio en ARS,
      código ACC-xx. Son "cajas de municiones", pero se venden por unidad como
      accesorios, no como el tipo regulado `municion`.
    - Fundas, linternas, cargadores, etc. (cotizan en USD) -> tipo `accesorios`,
      precio USD + equivalente ARS al TC, código ACCS-xx.

    Layout nuevo:  A=SKU B=MARCA C=MODELO(nombre) D=DESCRIPCION E=STOCK
                   H=EFECTIVO/TRANSFERENCIA  I=PVP TARJETA
                   J=TC 6 CUOTAS SIN INTERES (valor de CADA cuota, en ARS)
    Layout viejo:  A=nombre B=stock D=pvp.

    Tarjeta y 6 cuotas (I/J) vienen SIEMPRE en ARS, incluso para accesorios
    cotizados en USD (solo el efectivo/transferencia de esos está en USD).
    """
    ws = wb["Accesorios"]
    # El encabezado puede estar en la fila 1 o 2 según la versión del Excel.
    header_row = 1
    for hr in range(1, 6):
        vals = [txt(cell(ws, hr, c)).upper() for c in range(1, ws.max_column + 1)]
        if "MODELO" in vals or "SKU" in vals or "DESCRIPCION" in vals:
            header_row = hr
            break
    header = [txt(cell(ws, header_row, c)).upper() for c in range(1, ws.max_column + 1)]
    nuevo = "MODELO" in header  # encabezado del formato nuevo

    out = []
    idx_ammo = 0
    idx_acc = 0
    for r in range(header_row + 1, ws.max_row + 1):
        if nuevo:
            marca = txt(cell(ws, r, 2))
            name = txt(cell(ws, r, 3))
            stock = num(cell(ws, r, 5))
            col_efectivo = num(cell(ws, r, 8))  # H = EFECTIVO/TRANSFERENCIA
            col_tarjeta = num(cell(ws, r, 9))   # I = PVP TARJETA (ARS)
            col_cuota6 = num(cell(ws, r, 10))   # J = 6 CUOTAS SIN INTERES (ARS/cuota)
            if not name:
                continue
            # La munición 3DURBAN cotiza en ARS (H es el precio en pesos); los
            # accesorios cotizan en USD (H es el precio en dólares). Tarjeta (I) y
            # 6 cuotas (J) están SIEMPRE en ARS.
            up = name.upper()
            is_ammo = (marca.upper() == "3DURBAN"
                       or "CAJA" in up or "MUNIC" in up)
            precio_col = col_efectivo if col_efectivo is not None else col_tarjeta
        else:
            marca = ""
            name = txt(cell(ws, r, 1))
            stock = num(cell(ws, r, 2))
            col_tarjeta = None
            col_cuota6 = None
            if not name:
                continue
            # Layout viejo: solo munición.
            is_ammo = True
            precio_col = num(cell(ws, r, 4))
        if precio_col is None:
            continue
        stock_int = int(stock) if stock is not None else None

        if is_ammo:
            # 3DURBAN vende "cajas de municiones" cotizadas en ARS. Urban NO
            # llevó munición a la feria: estas cajas se venden como ACCESORIOS
            # (por unidad), NO como tipo `municion`. Se mantiene el código ACC-xx
            # para que un re-import sea idempotente si ya existían en la base.
            idx_ammo += 1
            precio_ars = round(precio_col, 2)
            out.append({
                "tipo": "accesorios",
                "marca": marca or "3DURBAN",
                "calibre": "",
                "modelo": "",
                "codigo": f"ACC-{idx_ammo:02d}",
                "descripcion": clean_desc(name),
                "precio_usd": 0,
                "stock": stock_int,
                "stock_inicial": stock_int,
                "efectivo_ars": precio_ars,
                "efectivo_usd": None,
                "tarjeta_ars": round(col_tarjeta, 2) if col_tarjeta else None,
                "cuota3_ars": None,
                "cuota6_ars": round(col_cuota6, 2) if col_cuota6 else None,
                "cuota12_ars": None,
            })
        else:
            # Accesorio cotizado en USD. Fiel al Excel = guardamos la referencia
            # en USD y su equivalente en ARS al TC del propio Excel (igual que las
            # armas Gral/Taurus). La app muestra estos montos, no recalcula.
            idx_acc += 1
            precio_usd = round(precio_col, 2)
            precio_ars = round(precio_col * TC, 2)
            out.append({
                "tipo": "accesorios",
                "marca": marca or "Accesorios",
                "calibre": "",
                "modelo": "",
                "codigo": f"ACCS-{idx_acc:02d}",
                "descripcion": clean_desc(name),
                "precio_usd": precio_usd,
                "stock": stock_int,
                "stock_inicial": stock_int,
                "efectivo_ars": precio_ars,
                "efectivo_usd": precio_usd,
                "tarjeta_ars": round(col_tarjeta, 2) if col_tarjeta else None,
                "cuota3_ars": None,
                "cuota6_ars": round(col_cuota6, 2) if col_cuota6 else None,
                "cuota12_ars": None,
            })
    total = idx_ammo + idx_acc
    if total:
        print(f"Accesorios emitidos: {total} "
              f"({idx_ammo} cajas 3DURBAN en ARS + {idx_acc} en USD)")
    return out


def read_tc(wb, default=1570.0):
    """Lee el TC del Excel (label 'TC:' en la cabecera de Gral, valor a la derecha)."""
    ws = wb["Gral"]
    for r in range(1, 5):
        for c in range(1, ws.max_column + 1):
            if txt(cell(ws, r, c)).upper().startswith("TC"):
                val = num(cell(ws, r, c + 1))
                if val:
                    return val
    return default


def _xlsx_cell(v):
    """Normaliza un valor para openpyxl: string vacío/espacios -> None."""
    if v is None:
        return None
    if isinstance(v, str) and v.strip() == "":
        return None
    return v


def make_dart_excel_compatible(path):
    """Reescribe los rels con Target absoluto (openpyxl escribe "/xl/...") a
    rutas relativas ("worksheets/sheet1.xml").

    El paquete Dart `excel` (^4.0.6) resuelve el worksheet como
    findFile('xl/' + Target); si Target es "/xl/worksheets/sheet1.xml" arma
    "xl//xl/..." → no lo encuentra y revienta con
    "Null check operator used on a null value". Los xlsx hechos con Excel usan
    Targets relativos, por eso esos sí importan; los de openpyxl no.
    """
    rels_name = "xl/_rels/workbook.xml.rels"
    with zipfile.ZipFile(path, "r") as zin:
        names = zin.namelist()
        if rels_name not in names:
            return
        data = {n: zin.read(n) for n in names}

    rels = data[rels_name].decode("utf-8")
    # Target="/xl/worksheets/sheet1.xml" -> "worksheets/sheet1.xml"
    fixed = re.sub(r'Target="/xl/', 'Target="', rels)
    # Cualquier otro Target absoluto "/algo" -> "algo"
    fixed = re.sub(r'Target="/', 'Target="', fixed)
    if fixed == rels:
        return
    data[rels_name] = fixed.encode("utf-8")

    tmp = path + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for n, content in data.items():
            zout.writestr(n, content)
    shutil.move(tmp, path)


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else \
        "/Users/abritos/Downloads/Catalogo_Armas_Nuevas_Urban_con_Cantidad.xlsx"
    dst = sys.argv[2] if len(sys.argv) > 2 else \
        "/Users/abritos/Downloads/urban_catalog_clean.xlsx"

    wb = openpyxl.load_workbook(src, data_only=True)
    global TC
    TC = read_tc(wb)
    print(f"TC del Excel: {TC}")
    products = []
    products += rows_gral_taurus(wb, "Gral", desc_col=5, price_lista_col=6)
    products += rows_gral_taurus(wb, "Taurus", desc_col=6, price_lista_col=5)
    products += rows_bersa(wb)
    products += rows_accesorios(wb)

    # Encabezados: canónicos + balas_por_caja (munición).
    headers = OUT_HEADERS + ["balas_por_caja"]

    out = openpyxl.Workbook()
    ws = out.active
    ws.title = "Catalogo"
    ws.append(headers)
    for p in products:
        # Celda vacía -> None (openpyxl la omite). Un string "" se escribiría
        # como <c t="inlineStr"/> autocerrado y el paquete Dart `excel` revienta
        # al buscar <t> ("Bad state: No element").
        ws.append([_xlsx_cell(p.get(h)) for h in headers])
    out.save(dst)
    # Deja el xlsx legible por el paquete Dart `excel` del importador.
    make_dart_excel_compatible(dst)

    # Resumen
    from collections import Counter
    by_type = Counter(p["tipo"] for p in products)
    by_marca = Counter(p["marca"] for p in products)
    print(f"OK -> {dst}")
    print(f"Total productos: {len(products)}")
    print("Por tipo:", dict(by_type))
    print("Marcas:", len(by_marca))
    # muestra
    print("\nEjemplos:")
    for p in products[:2] + products[-3:]:
        print(f"  [{p['tipo']}] {p['marca']} {p['modelo']} ({p['codigo']}) "
              f"cal={p['calibre']} stock={p['stock']} "
              f"efvo_ars={p['efectivo_ars']} usd={p['efectivo_usd']} "
              f"tarj={p['tarjeta_ars']} c3={p['cuota3_ars']}")


if __name__ == "__main__":
    main()
