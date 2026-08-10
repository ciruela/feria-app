#!/usr/bin/env python3
"""Normaliza la lista "PROMO GIRSAN" a un Excel canónico que el importador de la
app entiende, cargando la promo como PRECIOS FIJOS por producto.

Por qué precios fijos y no recargos globales:
  La promo "12 cuotas sin interés" aplica SOLO a Girsan. Los recargos de la app
  (`pricing_settings`) son globales por tenant, así que no sirven para acotar una
  promo a una marca. En cambio, cada producto puede llevar sus propios precios
  fijos (`FixedPrices`): cuando existen, `PricingService` los muestra tal cual y
  NO recalcula nada. Así la promo queda encapsulada en las 6 pistolas Girsan y no
  toca ni un peso del resto del catálogo.

Formato de entrada (hoja única, sin encabezados en A-H; encabezados solo en I-M):
  A codigo            | B marca            | C modelo/nombre
  D calibre           | E descripción larga| F precio USD (= efectivo)
  G categoría         | H stock            | I EFECTIVO/TRANSFERENCIA (USD)
  J PRECIO LISTA CON TARJETA (USD)         | K 3 SIN INTERES (ARS, valor de CADA cuota)
  L 6 SIN INTERES (ARS/cuota)              | M 12 SIN INTERES (ARS/cuota)

Mapeo a columnas canónicas de precios fijos:
  efectivo_usd = I                         (referencia dólar billete)
  efectivo_ars = I * TC                    (TC se deriva de la propia planilla)
  tarjeta_ars  = J * TC = K*3              (1 pago con tarjeta = total lista)
  cuota3_ars   = K,  cuota6_ars = L,  cuota12_ars = M
  precio_usd   = I                         (referencia USD; los fijos mandan)

Decisiones deliberadas (update de PRECIOS, mínimamente invasivo):
  - NO se emite `stock`  -> el importador conserva el stock actual del servidor
    (además la col. H trae 0 en algunas filas y las dejaría sin stock).
  - NO se emite `descripcion` -> conserva la descripción existente del producto.
  - El importador matchea por `codigo` (y, si no, por marca+modelo) y solo
    actualiza/crea; NUNCA borra productos ausentes del Excel. Por eso "actualizar
    solo esta lista" es seguro por diseño.

Uso:
  python3 scripts/girsan_promo_normalize.py \
      "~/Downloads/PROMO GIRSAN.xlsx" \
      ~/Downloads/girsan_promo_clean.xlsx
"""
import os
import re
import shutil
import sys
import zipfile

import openpyxl

OUT_HEADERS = [
    "tipo", "marca", "calibre", "modelo", "codigo",
    "precio_usd",
    "efectivo_usd", "efectivo_ars", "tarjeta_ars",
    "cuota3_ars", "cuota6_ars", "cuota12_ars",
]

# Índices (0-based) de las columnas de la planilla de entrada.
COL_CODIGO = 0   # A
COL_MARCA = 1    # B
COL_MODELO = 2   # C
COL_CALIBRE = 3  # D
COL_EFECTIVO_USD = 8   # I
COL_TARJETA_USD = 9    # J
COL_CUOTA3 = 10        # K
COL_CUOTA6 = 11        # L
COL_CUOTA12 = 12       # M

TIPO = "arma_corta"  # todas las Girsan de la lista son pistolas


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def r2(x):
    """Redondea a 2 decimales o None."""
    return None if x is None else round(x, 2)


def make_dart_excel_compatible(path):
    """Reescribe los rels con Target absoluto (openpyxl escribe "/xl/...") a
    rutas relativas ("worksheets/sheet1.xml").

    El paquete Dart `excel` resuelve el worksheet como findFile('xl/' + Target);
    si Target es "/xl/worksheets/sheet1.xml" arma "xl//xl/..." -> no lo encuentra
    y revienta con "Null check operator used on a null value". Los xlsx hechos con
    Excel usan Targets relativos, por eso esos sí importan; los de openpyxl no.
    """
    rels_name = "xl/_rels/workbook.xml.rels"
    with zipfile.ZipFile(path, "r") as zin:
        names = zin.namelist()
        if rels_name not in names:
            return
        data = {n: zin.read(n) for n in names}

    rels = data[rels_name].decode("utf-8")
    fixed = re.sub(r'Target="/xl/', 'Target="', rels)
    fixed = re.sub(r'Target="/', 'Target="', fixed)
    if fixed == rels:
        return
    data[rels_name] = fixed.encode("utf-8")

    tmp = path + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for n, content in data.items():
            zout.writestr(n, content)
    shutil.move(tmp, path)


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)

    src = os.path.expanduser(sys.argv[1])
    dst = os.path.expanduser(sys.argv[2])

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb.worksheets[0]

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Catalogo"
    out_ws.append(OUT_HEADERS)

    emitted = 0
    warnings = []

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        codigo = (str(row[COL_CODIGO]).strip() if row[COL_CODIGO] else "")
        marca = (str(row[COL_MARCA]).strip() if row[COL_MARCA] else "")
        modelo = (str(row[COL_MODELO]).strip() if row[COL_MODELO] else "")
        calibre = (str(row[COL_CALIBRE]).strip() if row[COL_CALIBRE] else "")

        efectivo_usd = num(row[COL_EFECTIVO_USD])
        tarjeta_usd = num(row[COL_TARJETA_USD])
        cuota3 = num(row[COL_CUOTA3])
        cuota6 = num(row[COL_CUOTA6])
        cuota12 = num(row[COL_CUOTA12])

        # Solo filas de producto reales (con código y precios de la promo).
        if not codigo or efectivo_usd is None or tarjeta_usd is None:
            continue
        if cuota3 is None:
            warnings.append(f"fila {i} ({codigo}): sin cuotas, se omite")
            continue

        # TC derivado de la propia planilla: total lista tarjeta = K*3 = J*TC.
        tc = (cuota3 * 3) / tarjeta_usd if tarjeta_usd else None
        if not tc:
            warnings.append(f"fila {i} ({codigo}): TC no derivable, se omite")
            continue

        efectivo_ars = efectivo_usd * tc
        tarjeta_ars = tarjeta_usd * tc  # 1 pago con tarjeta = total lista

        # Sanity: 3/6/12 sin interés => mismo total = lista tarjeta.
        totals = [c * n for c, n in ((cuota3, 3), (cuota6, 6), (cuota12, 12))
                  if c is not None]
        if any(abs(t - tarjeta_ars) > 1.0 for t in totals):
            warnings.append(
                f"fila {i} ({codigo}): cuotas no dan 'sin interés' exacto "
                f"(lista={tarjeta_ars:.0f}, totales={[round(t) for t in totals]})"
            )

        out_ws.append([
            TIPO, marca or "Girsan", calibre, modelo, codigo,
            r2(efectivo_usd),
            r2(efectivo_usd), r2(efectivo_ars), r2(tarjeta_ars),
            r2(cuota3), r2(cuota6), r2(cuota12),
        ])
        emitted += 1
        print(
            f"  {codigo:<18} ef=U$D{efectivo_usd:>6.0f} ({efectivo_ars:>12,.0f} ARS)"
            f" | lista U$D{tarjeta_usd:>6.0f} ({tarjeta_ars:>12,.0f} ARS)"
            f" | 12x {cuota12:>11,.2f}  TC={tc:.1f}"
        )

    out_wb.save(dst)
    # Deja el xlsx legible por el paquete Dart `excel` del importador.
    make_dart_excel_compatible(dst)
    print(f"\nEmitidas {emitted} pistolas Girsan -> {dst}")
    if warnings:
        print("\nAvisos:")
        for w in warnings:
            print(f"  - {w}")


if __name__ == "__main__":
    main()
